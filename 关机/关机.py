import os,PySimpleGUI as sg,openpyxl,datetime,random,sys,time
from PySimpleGUI import Button,Column, Frame,Radio,Window,Text,Input
from openpyxl.styles import Font, Alignment

class PopUpWindow():
    @classmethod
    def warning_window(cls, content, title='警告窗口'):
        sg.popup_ok(content,title='警告',keep_on_top = True)

    @classmethod
    def window_a(self, theme='GreenTan'):
        sg.theme(theme)
        col1 = Column([
            # Categories frame
            [Frame('操作种类:', [[
                Radio('关闭计算机', 'radio1', default=False, key='-SHUTDOWN-', size=(10, 1)),
                Radio('重启计算机', 'radio1', default=False, key='-RESTART-', size=(10, 1))
            ]], )]], pad=(0, 0))
        col3 = Column([[
            Frame('执行操作:', [[Column([[Button('确定'),
                                      Button('重置'),
                                      Button('取消'), ]], size=(234, 45), pad=(0, 0))]])
        ]], pad=(0, 0))
        col4 = Column([
            [Frame('延时执行时间(秒):', [[Text(), Column([[Input(key='-ACCOUNT-IN-', size=(15, 5))],
                                                   ], size=(220, 35), pad=(0, 0))]])], ], pad=(0, 0))

        layout = [[col1], [col4], [col3]]
        return layout



path = 'D://DATA//closingtime.xlsx'

if not os.path.exists(path):
    workbook = openpyxl.Workbook()
    worksheet = workbook[workbook.sheetnames[0]]
    worksheet.cell(row=1, column=1, value = '日期').alignment = Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=1, column=1, value = '日期').font = Font(bold=True)
    worksheet.cell(row=1, column=2, value = '时间').alignment = Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=1, column=2, value = '时间').font = Font(bold=True)
    workbook.save(path)

workbook = openpyxl.load_workbook(path)
worksheet = workbook[workbook.sheetnames[0]]

closing_day = datetime.datetime.strftime(datetime.datetime.now(),'%Y/%m/%d')
closing_time = datetime.datetime.strftime(datetime.datetime.now(),'%H:%M')
Time_from_work = datetime.datetime(int(datetime.datetime.strftime(datetime.datetime.now(),'%Y')),int(datetime.datetime.strftime(datetime.datetime.now(),'%m')),int(datetime.datetime.strftime(datetime.datetime.now(),'%d')),hour=17,minute=0,second=0)

if datetime.datetime.now() < Time_from_work:
    closing_time = '{}:{}'.format('17',random.randint(1,20))

flag = True
for i in range(1,worksheet.max_row+1):
    if worksheet.cell(row=i, column=1).value == closing_day:
        worksheet.cell(row=i, column=2,value=closing_time)
        flag = False
        break

if flag:
    worksheet.cell(row=worksheet.max_row+1, column=1, value=closing_day)
    worksheet.cell(row=worksheet.max_row, column=2, value=closing_time)
workbook.save(path)


while True:
    layout = PopUpWindow.window_a('Python')

    window = Window('电脑电源', layout, keep_on_top=True)

    event, values = window.read()
    # print(event, values)

    if event == sg.WIN_CLOSED:
        window.close()
        sys.exit(2)
        break
    elif event == '重置':
        window.close()
        continue
    elif event == '取消':
        sys.exit(2)
    elif event == '确定':
        # 判断是否选择了操作种类，如果'-SHUTDOWN-'和'-RESTART-'都没有选，则continue
        if values['-SHUTDOWN-'] == False and values['-RESTART-'] == False:
            window.close()
            continue
        window.close()
        # 判断延时时间是否有输入，是否全是数字。没有输入则设置为0，不全是数字则continue
        if values['-ACCOUNT-IN-'] == '':
            values['-ACCOUNT-IN-'] = '0'
        elif not values['-ACCOUNT-IN-'].isdigit():
            PopUpWindow.warning_window('延时时间必须是数字，不能是字母或特殊符号')
            window.close()
            continue
        break



time.sleep(0.5)

if values['-ACCOUNT-IN-'] == '' or values['-ACCOUNT-IN-'] == '0':
    shutdown_commond = "shutdown -s -t 1"
    restart_commond = "shutdown -r"
else:
    shutdown_commond = "shutdown -s -t " + values['-ACCOUNT-IN-']
    restart_commond = "shutdown -r -t " + values['-ACCOUNT-IN-']

if values['-SHUTDOWN-'] == True:
    if values['-ACCOUNT-IN-'] == '0':
        sg.popup('关闭所有电源',font=("",200),button_type=5,auto_close=True,auto_close_duration=5,background_color='#33FF00',text_color='#FF0000',button_color=('#6600CC','#FFFF33'),keep_on_top = True)
    # print(shutdown_commond)
    os.system(shutdown_commond)
    # os.system("shutdown -s -t 2")
else:
    # print(restart_commond)
    os.system(restart_commond)
    # os.system("shutdown -r")
